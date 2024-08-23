import tkinter as tk
from tkinter import scrolledtext, messagebox, BooleanVar, filedialog
from openpyxl import load_workbook
from db_function import *
import threading
import socket
import ipaddress
import re
import IPy
import pyperclip
import time
import configparser

# 创建配置解析器
config = configparser.ConfigParser()
# 读取配置文件
config.read('config.ini', encoding='utf-8')
whitelist_file = config.get('Settings', 'whitelist_file')
clipboard_interval = config.getint('Settings', 'clipboard_interval')

# 正则表达式，用于匹配IP地址和可能的端口号
ip_pattern = re.compile(r'\b(?:[0-9]{1,3}\.){3}[0-9]{1,3}(?::\d+)?\b')
# 正则表达式，用于匹配IPv6地址和可能的端口号
ipv6_regex = r'\b(?:[A-Fa-f0-9]{1,4}:){7}[A-Fa-f0-9]{1,4}\b|\b(?:[A-Fa-f0-9]{1,4}:)*:(?:[A-Fa-f0-9]{1,4}:)*[A-Fa-f0-9]{1,4}\b'
# 创建正则表达式对象
ipv6_pattern = re.compile(ipv6_regex)

db_name = "white_ips.db"

# 从文本中提取IP地址（支持IPv4和IPv6）
def extract_ips_ipv6(text):
    ips = ipv6_pattern.findall(text)
    valid_ips = []
    # 过滤出有效的IPv6地址
    for ip in ips:
        try:
            if ipaddress.ip_address(ip):
                valid_ips.append(ip)
        except:
            print(f"{ip}解析失败")
    return valid_ips


# 从文本中提取IP地址
def extract_ips(text):
    ips = ip_pattern.findall(text)
    return [ip for ip in ips if ipaddress.ip_address(ip)]


def read_whitelist_from_db():
    whitelist_ipv4 = set()
    whitelist_ipv6 = set()
    whitelist_ipv6_networks = set()
    ip_list = query_all_ips(db_name)
    for line in ip_list:
        line = line.strip()
        if not line:  # 如果行是空的，跳过
            continue
        # 处理IP范围、IP掩码和单个IP，与之前相同
        try:
            ip_network = ipaddress.ip_network(line, strict=False)
        except:
            print(f"{line}不是一个合法的IP地址或网段")
        try:
            if '/' in line:
                if isinstance(ip_network, ipaddress.IPv6Network):
                    whitelist_ipv6_networks.add(line)  # 添加IPv6网段
                else:
                    ip = IPy.IP(line)
                    for x in ip:
                        whitelist_ipv4.add(str(x))
            else:
                if isinstance(ip_network, ipaddress.IPv6Network):
                    whitelist_ipv6.add(line)
                else:
                    whitelist_ipv4.add(line)  # 直接添加到集合中
        except:
            messagebox.showerror("警告", f"加载{line}白名单失败，请检查白名单地址格式")
    return whitelist_ipv4, whitelist_ipv6, whitelist_ipv6_networks

# 读取加白地址并转换为IP网络列表
def read_whitelist(file_path):
    whitelist_ipv4 = set()
    whitelist_ipv6 = set()
    whitelist_ipv6_networks = set()
    with open(file_path, 'r') as file:
        for line in file:
            line = line.strip()
            if not line:  # 如果行是空的，跳过
                continue
            # 处理IP范围、IP掩码和单个IP，与之前相同
            ip_network = ipaddress.ip_network(line, strict=False)
            try:
                if '/' in line:
                    if isinstance(ip_network, ipaddress.IPv6Network):
                        whitelist_ipv6_networks.add(line)  # 添加IPv6网段
                    else:
                        ip = IPy.IP(line)
                        for x in ip:
                            whitelist_ipv4.add(str(x))
                else:
                    if isinstance(ip_network, ipaddress.IPv6Network):
                        whitelist_ipv6.add(line)
                    else:
                        whitelist_ipv4.add(line)  # 直接添加到集合中
            except:
                messagebox.showerror("警告", f"加载{line}白名单失败，请检查白名单地址格式")
    return whitelist_ipv4, whitelist_ipv6, whitelist_ipv6_networks


def is_ip_in_ipv6_network(ip, network_cidr):
    try:
        # 将IPv6地址转换为IPv6Address对象
        ip_address = ipaddress.IPv6Address(ip)
        # 将网段的CIDR表示法转换为IPv6Network对象
        ip_network = ipaddress.ip_network(network_cidr, strict=False)
        print(f"正在检测{ip}")
    except ValueError as e:
        print(f"{ip}检测失败,{e}")
        return False
    # 检查IPv6地址是否在IPv6网段内
    return ip_address in ip_network


# 剔除加白地址
def filter_ips(input_ips, whitelist):
    return [ip for ip in input_ips if ip not in whitelist]


def filter_ips_ipv6(input_ips, whitelist_ipv6, whitelist_ipv6_networks):
    filtered_ips = []  # 使用列表来存储不在白名单中的IP
    for ip in input_ips:
        if ip in whitelist_ipv6:
            continue  # 如果IP在白名单中，跳过
        ip_in_network = False
        for network in whitelist_ipv6_networks:
            if is_ip_in_ipv6_network(ip, network):
                ip_in_network = True
                break
        if not ip_in_network:  # 如果IP不在任何网络中，添加到结果列表
            filtered_ips.append(ip)
    return filtered_ips


def unique_ips(ip_list):
    return list(set(ip_list))


# 更新输出框
def update_output(ip_list, unique=False):
    output_area.delete(1.0, tk.END)
    output_area.insert(tk.INSERT, '\n'.join(ip_list))
    if copy_to_clipboard_var.get():
        copy_to_clipboard()
    if unique:
        output_ips = unique_ips(ip_list)
        output_area.delete(1.0, tk.END)
        output_area.insert(tk.INSERT, '\n'.join(output_ips))


def update_input(ip_list):
    input_area.delete(1.0, tk.END)
    input_area.insert(tk.INSERT, '\n'.join(ip_list))


# 切换自动去重开关
def toggle_unique():
    global unique_enabled
    unique_enabled = not unique_enabled
    if unique_enabled:
        aoto_unique_button.config(text="关闭自动去重")
        # 当开启去重时，立即去重当前输出框的IP
        ip_list = output_area.get(1.0, tk.END).strip().split('\n')
        update_output(ip_list, unique=True)
    else:
        aoto_unique_button.config(text="开启自动去重")
        # 当关闭去重时，显示输出框原始IP（可能包含重复）
        update_output(output_area.get(1.0, tk.END).strip().split('\n'))


def ip_unique():
    ip_list = output_area.get(1.0, tk.END).strip().split('\n')
    output_ips = unique_ips(ip_list)
    output_area.delete(1.0, tk.END)
    output_area.insert(tk.INSERT, '\n'.join(output_ips))
    update_output_count_label()  # 更新输出框IP数量标签


# 复制到剪贴板
def copy_to_clipboard():
    try:
        root.clipboard_clear()
        root.clipboard_append(output_area.get(1.0, tk.END))
    except Exception as e:
        messagebox.showerror("复制失败", f"复制时发生错误: {e}")


# 过滤按钮的回调函数
def on_filter():
    input_text = input_area.get(1.0, tk.END)
    input_ips_ipv4 = extract_ips(input_text)
    filtered_ips_ipv4 = filter_ips(input_ips_ipv4, whitelist)
    # print(filtered_ips_ipv4)
    input_ips_ipv6 = extract_ips_ipv6(input_text)
    filtered_ips_ipv6 = filter_ips_ipv6(input_ips_ipv6, whitelist_ipv6, whitelist_ipv6_networks)
    input_ips = input_ips_ipv4 + input_ips_ipv6
    update_input(input_ips)
    # print("IPV6:")
    # print(filtered_ips_ipv6)
    filtered_ips = filtered_ips_ipv6 + filtered_ips_ipv4
    print(f"filed:{filtered_ips}")
    update_output(filtered_ips, unique=unique_enabled)
    update_output_count_label()


# 创建复制开关
def toggle_copy_to_clipboard():
    copy_to_clipboard_var.set(not copy_to_clipboard_var.get())
    if copy_to_clipboard_var.get():
        auto_copy_button.config(text="关闭自动复制")
    else:
        auto_copy_button.config(text="开启自动复制")


# 创建文件选择按钮
def read_ips_from_excel(file_path):
    # 加载Excel文件
    wb = load_workbook(file_path)
    # 选择工作表
    ws = wb.active  # 或者 wb['Sheet1'] 来指定工作表
    # 读取第一列的数据
    column_data = [str(ws[f'A{i}'].value) for i in range(1, ws.max_row + 1)]
    return [data for data in column_data]


def select_file():
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx;*.xls;*.csv")])
    if file_path:
        ips = read_ips_from_excel(file_path)
        text = '\n'.join(ips)
        ips_ipv4 = extract_ips(text)
        ips_ipv6 = extract_ips_ipv6(text)
        filted_ips = ips_ipv6 + ips_ipv4
        input_area.delete(1.0, tk.END)
        input_area.insert(tk.INSERT, '\n'.join(filted_ips))
        update_ip_count_label()


# 居中窗口
def center_window(width=800, height=400):
    # 获取屏幕尺寸
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()
    # 计算x和y坐标
    x = (screen_width - width) // 2
    y = (screen_height - height) // 2
    root.geometry(f'{width}x{height}+{x}+{y}')


def monitor_clipboard():
    global monitor_enabled
    while True:
        with lock:
            if monitor_enabled:
                text = pyperclip.paste()
                print(f"clpblod:{text}")
                input_ips = extract_ips(text)
                print(f" input_ips list is {input_ips}")
                input_ips_ipv6 = extract_ips_ipv6(text)
                filtered_ips_ipv4 = filter_ips(input_ips, whitelist)
                filtered_ips_ipv6 = filter_ips_ipv6(input_ips_ipv6, whitelist_ipv6, whitelist_ipv6_networks)
                ip_list = filtered_ips_ipv4 + filtered_ips_ipv6
                print(f" ip list is {ip_list}")
                filtered_text = '\n'.join(ip_list)
                if filtered_text != text:
                    pyperclip.copy(filtered_text)
                    print(f"fi is {filtered_text}")
                    status_label.config(text="剪贴板检测到IP，已自动更新过滤ip.")
            clipboard_interval = config.getint('Settings', 'clipboard_interval')
            time.sleep(clipboard_interval)  # 暂停一段时间再检查


def on_monitor_toggle():
    global monitor_enabled
    if monitor_enabled:
        monitor_enabled = False
        status_label.config(text="监控粘贴板关闭中")
    else:
        monitor_enabled = True
        status_label.config(text="监控粘贴板已开启")
        threading.Thread(target=monitor_clipboard, daemon=True).start()


# 更新输入框IP数量标签
def update_ip_count_label():
    text = input_area.get(1.0, tk.END)
    ip_count = len(text.strip().split('\n'))
    # print(text.strip().split('\n'))
    ip_count_label.config(text=f"输入IP 数量: {ip_count}")


# 更新输出框IP数量标签
def update_output_count_label():
    text = output_area.get(1.0, tk.END)
    lines = text.strip().split('\n')
    clean_lines = [line.rstrip() for line in lines]
    out_ip_count = len(clean_lines)
    # print(text.strip().split('\n'))
    if text.strip().split('\n')[0] != '':
        output_ip_count_label.config(text=f"输出IP数量: {out_ip_count}")
    else:
        output_ip_count_label.config(text=f"输出IP数量: 0")


def sort_ips_ipv6(ipv6_addresses):
    # 使用ipaddress.IPv6Address将字符串转换为IPv6地址对象
    ipv6_objects = [ipaddress.IPv6Address(addr) for addr in ipv6_addresses]
    # 对IPv6地址对象进行排序
    sorted_ipv6_objects = sorted(ipv6_objects)
    # 将排序后的IPv6地址对象转换回字符串，如果需要的话
    sorted_ipv6_addresses = [str(addr) for addr in sorted_ipv6_objects]
    return sorted_ipv6_addresses


def is_valid_ipv6(ip):
    try:
        ipaddress.IPv6Address(ip)
        return True
    except:
        return False


def sort_ips(ips_list):
    ips_ipv4 = []
    ips_ipv6 = []
    for ip in ips_list:
        if is_valid_ipv6(ip):
            ips_ipv6.append(ip)
        else:
            ips_ipv4.append(ip)
    sorted_ips_ipv6 = sort_ips_ipv6(ips_ipv6)
    sorted_ips_ipv4 = sorted(ips_ipv4, key=socket.inet_aton)
    return sorted_ips_ipv4, sorted_ips_ipv6


def on_sort_ips():
    text_in = input_area.get(1.0, tk.END)
    text_out = output_area.get(1.0, tk.END)
    in_ips = text_in.strip().split('\n')
    out_ips = text_out.strip().split('\n')

    sorted_in_ips_ipv4, sorted_in_ips_ipv6 = sort_ips(in_ips)
    sorted_in_ips = sorted_in_ips_ipv6 + sorted_in_ips_ipv4

    sorted_out_ips_ipv4, sorted_out_ips_ipv6 = sort_ips(out_ips)
    sorted_out_ips = sorted_out_ips_ipv6 + sorted_out_ips_ipv4

    output_area.delete(1.0, tk.END)
    output_area.insert(tk.INSERT, '\n'.join(sorted_out_ips))
    input_area.delete(1.0, tk.END)
    input_area.insert(tk.INSERT, '\n'.join(sorted_in_ips))
    update_output_count_label()
    update_ip_count_label()
    cdns = [x for x in sorted_in_ips if x not in sorted_out_ips]
    print(cdns)
    for cdn in cdns:
        index = sorted_in_ips.index(cdn) + 1
        input_area.tag_configure("red", foreground="red")
        input_area.tag_add("red", f"{index}.0", f"{index}.end")


# 绑定输入框文本变化事件，更新IP数量标签
def on_input_text_change(event):
    update_ip_count_label()


# 绑定输出框文本变化事件，更新IP数量标签
def on_output_text_change(event):
    update_output_count_label()


# 清空输入输出框并重置IP计数标签
def clear_ip_boxes():
    input_area.delete(1.0, tk.END)
    output_area.delete(1.0, tk.END)
    output_ip_count_label.config(text=f"输出IP数量: 0")
    ip_count_label.config(text=f"输入IP 数量: 0")


def add_ips_to_whitelist(new_ips_list, file_path):
    # 创建一个集合来存储所有唯一的 IP 地址
    unique_ips = set(new_ips_list)
    # 打开文件，读取现有的 IP 地址
    with open(file_path, 'r') as file:
        for line in file:
            ip = line.strip()
            if ip:  # 确保不添加空行
                unique_ips.add(ip)
    with open(file_path, 'w') as file:
        for ip in unique_ips:
            file.write(ip + '\n')


def add_to_whitelist():
    # 创建一个顶级窗口作为输入框
    top = tk.Toplevel(root)
    top.title("添加白名单 IP")

    # 定义一个函数，用于在顶级窗口中添加 IP 并关闭窗口
    def add_ips_and_close():
        # 从 ScrolledText 获取 IP 列表
        ip_text = st.get(1.0, tk.END)
        new_ips_list = [ip.strip() for ip in ip_text.split('\n') if ip.strip()]
        if new_ips_list:  # 如果 IP 列表不为空
            file_path = config.get('Settings', 'whitelist_file')
            # 去重添加到文件和内存中的白名单
            add_ips_to_whitelist(new_ips_list, file_path)
            global whitelist, whitelist_ipv6, whitelist_ipv6_networks
            whitelist, whitelist_ipv6, whitelist_ipv6_networks = read_whitelist(file_path)
        messagebox.showinfo("成功", "IP 地址已添加到白名单。")
        top.destroy()  # 关闭顶级窗口

    # 创建 ScrolledText 组件
    st = scrolledtext.ScrolledText(top, width=50, height=10)
    st.pack(fill=tk.BOTH, expand=True)
    # 创建一个按钮来添加 IP 并关闭窗口
    add_button = tk.Button(top, text="添加 IP 并关闭", command=add_ips_and_close)
    add_button.pack(side=tk.BOTTOM)
    # 聚焦于 ScrolledText 组件
    st.focus_set()


def read_whitelist_from_txt():
    file_path = filedialog.askopenfilename(filetypes=[("Text files", "*.txt")])
    #txt_file = 'ips.txt'  # 假设IP列表存储在ips.txt文件中
    count = 0
    with open(file_path, 'r') as file:
        for line in file:
            ip = line.strip()  # 移除行尾的换行符
            if ip:  # 如果行不为空
                if insert_ip(db_name,ip):
                    count = count + 1
    messagebox.showinfo("成功", f"{count}条地址已添加到白名单。")
    refresh_whitelist_in_memory_from_db()
    whitelist_operations_panel.destroy()


def find_ip_in_line(text):
    # 正则表达式，用于匹配IP地址和可能的端口号
    #ip_pattern = re.compile(r'\b(?:[0-9]{1,3}\.){3}[0-9]{1,3}(?::\d+)?\b')
    ip_pattern = re.compile(r'\b(?:[0-9]{1,3}\.){3}[0-9]{1,3}(?:/\d+)?\b')

    # # 正则表达式，用于匹配IPv6地址和可能的端口号
    # ipv6_regex = re.compile(r'\b(?:[A-Fa-f0-9]{1,4}:){7}[A-Fa-f0-9]{1,4}\b|'
    #                         r'\b(?:[A-Fa-f0-9]{1,4}:)*:(?:[A-Fa-f0-9]{1,4}:)*[A-Fa-f0-9]{1,4}\b')
    # 正则表达式，用于匹配IPv6地址和可能的CIDR网段
    ipv6_regex = re.compile(
        r'\b(?:(?:[A-Fa-f0-9]{1,4}:){7}(?:[A-Fa-f0-9]{1,4}|:)|'  # 完整地址或以冒号结束
        r'(?:[A-Fa-f0-9]{1,4}:){6}(?:[A-Fa-f0-9]{1,4}|:[A-Fa-f0-9]{1,4}|:)|'  # 省略最后一个段
        r'(?:[A-Fa-f0-9]{1,4}:){5}(?::[A-Fa-f0-9]{1,4}){1,2}|'  # 省略最后两个段
        r'(?:[A-Fa-f0-9]{1,4}:){4}(?::[A-Fa-f0-9]{1,4}){1,3}|'  # 省略最后三个段
        r'(?:[A-Fa-f0-9]{1,4}:){3}(?::[A-Fa-f0-9]{1,4}){1,4}|'  # 省略最后四个段
        r'(?:[A-Fa-f0-9]{1,4}:){2}(?::[A-Fa-f0-9]{1,4}){1,5}|'  # 省略最后五个段
        r'(?:[A-Fa-f0-9]{1,4}:){1}(?::[A-Fa-f0-9]{1,4}){1,6}|'  # 省略最后六个段
        r':(?::[A-Fa-f0-9]{1,4}){1,7}|'  # 只有冒号开头，后面是七个段
        r'fe80:(?::[A-Fa-f0-9]{0,4}){0,4}%[0-9a-zA-Z]{1,}|'  # 特殊地址fe80::/10，可能带有zone index
        r'::(?:ffff(?::0{1,4}){0,1}:){0,1}(?:(?:25[0-5]|(2[0-4]|1{0,1}[0-9]){0,1}[0-9])\.){3}(?:25[0-5]|(2[0-4]|1{0,1}[0-9]){0,1}[0-9])\b)'
        # IPv4映射地址
    )
    ipv4_match = ip_pattern.search(text)
    # 如果IPv4匹配成功，则返回结果
    if ipv4_match:
        return ipv4_match.group()
    # 否则尝试匹配IPv6地址和端口号
    ipv6_match = ipv6_regex.search(text)
    # 如果IPv6匹配成功，则返回结果
    if ipv6_match:
        return ipv6_match.group()
    # 如果都没有匹配到，返回None
    return None

def is_valid_ip(ip):
    try:
        # 尝试将输入解析为IPv4或IPv6地址或网段
        ip_obj = ipaddress.ip_interface(ip).ip
        # 如果ip_obj是有效的IPv4或IPv6地址或网段，返回True
        return isinstance(ip_obj, (ipaddress.IPv4Address, ipaddress.IPv6Address))
    except ValueError:
        # 如果解析失败，返回False
        return False

def read_whitelist_from_excel():
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx;*.xls;*.csv")])
    # 加载Excel工作簿
    wb = load_workbook(file_path)
    # 假设数据在第一个工作表上
    ws = wb.active
    # 遍历工作表中的行（跳过标题行）
    count = 0
    for row in ws.iter_rows(min_row=1, values_only=True):
        ip = None
        description = None
        try:
            ip, description = row
        except:
            ip = row[0]
            description = '无描述'
        finally:
            if not description:
                description = '无描述'
        if is_valid_ip(ip):
            if insert_single_data(db_name,ip,description):
                count = count + 1
    messagebox.showinfo("成功", f"{count}条地址已添加到白名单。")
    refresh_whitelist_in_memory_from_db()
    global whitelist_operations_panel
    whitelist_operations_panel.destroy()

def view_whitelist_ips():
    popup = tk.Toplevel()
    popup.title('Data Display Popup')

    # 设置弹窗的固定宽度和高度
    popup_width = 400
    popup_height = 300

    # 获取屏幕宽度和高度，以计算弹窗的居中位置
    screen_width = popup.winfo_screenwidth()
    screen_height = popup.winfo_screenheight()

    # 计算弹窗的X和Y坐标，使其居中
    x = (screen_width // 2) - (popup_width // 2)
    y = (screen_height // 2) - (popup_height // 2)

    # 设置弹窗的几何形状（位置和大小）
    popup.geometry(f'{popup_width}x{popup_height}+{x}+{y}')

    # 创建一个滚动文本框，并让它填满弹窗
    text_area = scrolledtext.ScrolledText(popup, wrap=tk.WORD)
    text_area.pack(fill=tk.BOTH, expand=True)  # 填满并扩展

    rows = query_all_ip_and_descs(db_name)
    text = 'IP            | 描述\n'
    for row in rows:
        print(f'row is {row}')
        text = text + f'{row[0]} , {row[1]}\n'

    text_area.insert(tk.END, text)



def export_whitelist_ips():
    # 您的实现代码
    export_to_csv(db_name,"whitelist.csv")
    messagebox.showinfo("成功", "白名单已导出到whitelist.csv。")
    global whitelist_operations_panel
    whitelist_operations_panel.destroy()

def delete_ip_from_whitelist():
    top = tk.Toplevel(root)
    top.title("在下方输入IP，将批量从白名单中删除")
    # 定义一个函数，用于在顶级窗口中添加 IP 并关闭窗口
    def delete_ips_and_close():
        # 从 ScrolledText 获取 IP 列表
        ip_text = st.get(1.0, tk.END)
        del_ips_list = [ip.strip() for ip in ip_text.split('\n') if ip.strip()]
        count = 0
        if del_ips_list:  # 如果 IP 列表不为空
            for ip in del_ips_list:
                if delete_data_by_ip(db_name,ip):
                    count = count + 1
        messagebox.showinfo("成功", f"{count}个IP已从白名单中删除")
        refresh_whitelist_in_memory_from_db()
        top.destroy()  # 关闭顶级窗口
        global whitelist_operations_panel
        whitelist_operations_panel.destroy()
    # 创建 ScrolledText 组件
    st = scrolledtext.ScrolledText(top, width=50, height=10)
    st.pack(fill=tk.BOTH, expand=True)
    # 创建一个按钮来添加 IP 并关闭窗口
    add_button = tk.Button(top, text="删除IP", command=delete_ips_and_close)
    add_button.pack(side=tk.BOTTOM)
    # 聚焦于 ScrolledText 组件
    st.focus_set()


def clear_whitelist_ips():
    clear_table_data(db_name)
    messagebox.showinfo("成功", "白名单数据已全部删除")
    global whitelist_operations_panel
    whitelist_operations_panel.destroy()
    refresh_whitelist_in_memory_from_db()


def read_whitelist_from_text_box():
    top = tk.Toplevel(root)
    top.title("在下方输入IP，将批量从添加到白名单中")
    # 定义一个函数，用于在顶级窗口中添加 IP 并关闭窗口
    def insert_ips_and_close():
        # 从 ScrolledText 获取 IP 列表
        ip_text = st.get(1.0, tk.END)
        ins_ips_list = [ip.strip() for ip in ip_text.split('\n') if ip.strip()]
        count = 0
        if ins_ips_list:  # 如果 IP 列表不为空
            for ip in ins_ips_list:
                if insert_ip(db_name,ip):
                    count = count + 1
        messagebox.showinfo("成功", f"{count}个IP已添加到白名单中")
        refresh_whitelist_in_memory_from_db()
        top.destroy()  # 关闭顶级窗口
        global whitelist_operations_panel
        whitelist_operations_panel.destroy()
    # 创建 ScrolledText 组件
    st = scrolledtext.ScrolledText(top, width=50, height=10)
    st.pack(fill=tk.BOTH, expand=True)
    # 创建一个按钮来添加 IP 并关闭窗口
    add_button = tk.Button(top, text="添加IP", command=insert_ips_and_close)
    add_button.pack(side=tk.BOTTOM)
    # 聚焦于 ScrolledText 组件
    st.focus_set()


def open_whitelist_operations_panel():
    global whitelist_operations_panel
    whitelist_operations_panel = tk.Toplevel(root)
    whitelist_operations_panel.title("白名单IP操作")

    # 居中窗口
    panel_width = 200
    panel_height = 340
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()
    x = (screen_width - panel_width) // 2
    y = (screen_height - panel_height) // 2
    whitelist_operations_panel.geometry(f'{panel_width}x{panel_height}+{x}+{y}')

    # 从输入框白名单IP按钮
    load_button = tk.Button(whitelist_operations_panel, text="插入白名单IP",
                                command=lambda: read_whitelist_from_text_box())
    load_button.pack(fill=tk.BOTH, expand=True, pady=5)

    # 从txt文件插入白名单IP按钮
    load_txt_button = tk.Button(whitelist_operations_panel, text="从txt插入白名单IP",
                                command=lambda: read_whitelist_from_txt())
    load_txt_button.pack(fill=tk.BOTH, expand=True,pady=5)

    # 从Excel插入白名单IP按钮
    load_excel_button = tk.Button(whitelist_operations_panel, text="从Excel插入白名单IP",
                                  command=lambda: read_whitelist_from_excel())
    load_excel_button.pack(fill=tk.BOTH, expand=True,pady=5)

    # 查看白名单IP按钮
    view_whitelist_button = tk.Button(whitelist_operations_panel, text="查看白名单IP", command=view_whitelist_ips)
    view_whitelist_button.pack(fill=tk.BOTH, expand=True, pady=5)

    # 导出白名单IP按钮
    export_whitelist_button = tk.Button(whitelist_operations_panel, text="导出白名单IP", command=export_whitelist_ips)
    export_whitelist_button.pack(fill=tk.BOTH, expand=True,pady=5)

    # 根据IP删除白名单IP按钮
    delete_ip_button = tk.Button(whitelist_operations_panel, text="根据IP删除白名单IP",
                                 command=lambda: delete_ip_from_whitelist())
    delete_ip_button.pack(pady=5,fill=tk.BOTH, expand=True)

    # 清除所有白名单IP按钮
    clear_whitelist_button = tk.Button(whitelist_operations_panel, text="清除所有白名单IP", command=clear_whitelist_ips)
    clear_whitelist_button.pack(fill=tk.BOTH, expand=True, pady=5)


def refresh_whitelist_in_memory_from_db():
    global whitelist, whitelist_ipv6, whitelist_ipv6_networks
    whitelist, whitelist_ipv6, whitelist_ipv6_networks = read_whitelist_from_db()


def init_db(init_db_name):
    create_table(init_db_name)

global status_label
monitor_enabled = False
lock = threading.Lock()
init_db(db_name)
# 主窗口
root = tk.Tk()
root.title("IP Filter V3.1")

# 配置文件路径
file_path = config.get('Settings', 'whitelist_file')
whitelist, whitelist_ipv6, whitelist_ipv6_networks = read_whitelist_from_db()
#print(f"white_list is{whitelist}")
# 居中窗口
center_window(730, 430)

# 输入框和输出框的创建与之前相同，这里省略以节省空间
# 输入框
input_frame = tk.Frame(root)
input_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
input_area = scrolledtext.ScrolledText(input_frame, width=30, height=10)
input_area.pack(fill=tk.BOTH, expand=True)

# 输出框
output_frame = tk.Frame(root)
output_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)
output_area = scrolledtext.ScrolledText(output_frame, width=30, height=10)
output_area.pack(fill=tk.BOTH, expand=True)

# 创建过滤按钮
filter_button = tk.Button(root, text="Filter IPs", command=on_filter)
filter_button.pack(side=tk.BOTTOM, fill=tk.X)
# 创建清空按钮
clear_button = tk.Button(root, text="清空IP", command=clear_ip_boxes)
clear_button.pack(side=tk.BOTTOM, fill=tk.X)
empty_space1 = tk.Label(root, text="")
empty_space1.pack(side=tk.BOTTOM, pady=1)

# 创建复制开关按钮
copy_to_clipboard_var = BooleanVar(value=True)
auto_copy_button = tk.Button(root, text="关闭自动复制", command=toggle_copy_to_clipboard)
auto_copy_button.pack(side=tk.BOTTOM, fill=tk.X)
# 创建复制按钮
copy_button = tk.Button(root, text="复制到剪贴板", command=copy_to_clipboard)
copy_button.pack(side=tk.BOTTOM, fill=tk.X)
empty_space2 = tk.Label(root, text="")
empty_space2.pack(side=tk.BOTTOM, pady=1)

# 创建自动去重开关按钮
unique_enabled = False
aoto_unique_button = tk.Button(root, text="开启自动去重", command=toggle_unique)
aoto_unique_button.pack(side=tk.BOTTOM, fill=tk.X)
# 去重按钮
unique_button = tk.Button(root, text="输出IP去重", command=ip_unique)
unique_button.pack(side=tk.BOTTOM, fill=tk.X)
empty_space3 = tk.Label(root, text="")
empty_space3.pack(side=tk.BOTTOM, pady=1)

# IP排序
sort_input_button = tk.Button(root, text="排序IP并高亮", command=on_sort_ips)
sort_input_button.pack(side=tk.BOTTOM, fill=tk.X)
# 从文件读取
file_select_button = tk.Button(root, text="从Excel文件中读取IP", command=select_file)
file_select_button.pack(side=tk.BOTTOM, fill=tk.X)
# 添加白名单按钮
# add_whitelist_button = tk.Button(root, text="添加白名单", command=add_to_whitelist)
# add_whitelist_button.pack(side=tk.BOTTOM, fill=tk.X)

# 创建操作白名单IP按钮
whitelist_operations_panel = None
whitelist_operations_button = tk.Button(root, text="操作白名单IP", command=open_whitelist_operations_panel)
whitelist_operations_button.pack(side=tk.BOTTOM, fill=tk.X)

# 创建输入框IP数量标签
ip_count_label = tk.Label(input_frame, text="输入IP 数量: 0")
ip_count_label.pack(side=tk.BOTTOM)
input_area.bind('<KeyRelease>', on_input_text_change)
# 创建输出框IP数量标签
output_ip_count_label = tk.Label(output_frame, text="输出IP数量: 0")
output_ip_count_label.pack(side=tk.BOTTOM)
output_area.bind('<KeyRelease>', on_output_text_change)

# 创建一个按钮来切换监控状态
toggle_button = tk.Button(root, text="监控剪切板", command=on_monitor_toggle)
toggle_button.pack()
# 创建一个标签来显示状态信息
status_label = tk.Label(root, text="监控粘贴板已关闭")
status_label.pack()

try:
    root.mainloop()
except Exception as e:
    # 这里可以处理异常，例如记录日志或显示错误消息
    print(f"发生未处理的异常: {e}")
    # 根据需要，可以添加 messagebox 来显示错误信息
    messagebox.showerror("错误", f"发生未处理的异常: {e}")
finally:
    # 这里可以执行一些清理工作，无论是否发生异常都会执行
    pass
